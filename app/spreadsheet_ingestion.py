"""Spreadsheet (.xlsx) input: an attached workbook is converted server-side
into a bounded text table per sheet and folded into the SAME plain-text
document-attachment path a .txt/.csv file already goes through — see
app/audio_ingestion.py for the identical "convert to something the provider
already understands, before it ever reaches the provider or persistence
layer" pattern this mirrors. providers.py never sees the spreadsheetml
mime; by the time `req.files` reaches the orchestrator, an .xlsx attachment
looks exactly like a plain-text one.

CSV needs no equivalent module: a .csv file is already plain text, so the
frontend just normalizes its mime to `text/plain` before sending it
(matching how it already treats an unrecognized/.md mime) and the existing
text/plain path handles it with zero backend changes — see App.tsx's
`isDocumentFile`/mime-normalization in `handleFilesSelected`.
"""

from __future__ import annotations

import base64
import binascii
import csv
import io
from io import BytesIO

from fastapi import HTTPException
from openpyxl import load_workbook

from .schemas import _XLSX_MIME, FileAttachment
from .telemetry import logger

# A bound per sheet, not per workbook — a wide/tall spreadsheet is capped
# rather than rejected, with an explicit note in the injected content so the
# model knows data was cut rather than silently seeing a partial table.
_MAX_ROWS_PER_SHEET = 200
_MAX_COLS_PER_SHEET = 50

_XLSX_DATA_URL_PREFIX = f"data:{_XLSX_MIME};base64,"

# Preview grid bounds for POST /v1/spreadsheet-preview (app/routers/media.py)
# — a much tighter cap than the text-extraction bounds above, since this is
# an inline UI preview (first sheet only, a glance at the shape of the data),
# not the full multi-sheet table folded into a model's context.
_PREVIEW_MAX_ROWS = 50
_PREVIEW_MAX_COLS = 20


def xlsx_preview_rows(raw: bytes) -> tuple[list[list[str]], int, int]:
    """(rows, total_rows, total_cols) for the FIRST sheet of `raw` (an .xlsx
    workbook), capped at _PREVIEW_MAX_ROWS x _PREVIEW_MAX_COLS — the inline
    preview grid for a generated spreadsheet, distinct from xlsx_to_text's
    full multi-sheet extraction (that one feeds a model's context; this one
    feeds a UI table). Raises ValueError for anything openpyxl can't parse,
    same broad catch as xlsx_to_text (see its docstring)."""
    try:
        workbook = load_workbook(BytesIO(raw), data_only=True, read_only=True)
    except Exception as err:
        raise ValueError("not a valid .xlsx file") from err
    try:
        sheet = workbook.worksheets[0] if workbook.worksheets else None
        if sheet is None:
            return [], 0, 0
        total_rows = sheet.max_row or 0
        total_cols = sheet.max_column or 0
        shown_rows = min(total_rows, _PREVIEW_MAX_ROWS)
        shown_cols = min(total_cols, _PREVIEW_MAX_COLS)
        rows: list[list[str]] = []
        if shown_rows > 0 and shown_cols > 0:
            for row in sheet.iter_rows(
                min_row=1, max_row=shown_rows, max_col=shown_cols, values_only=True
            ):
                rows.append([_cell_text(v) for v in row])
        return rows, total_rows, total_cols
    finally:
        workbook.close()


def csv_preview_rows(raw: bytes) -> tuple[list[list[str]], int, int]:
    """Same (rows, total_rows, total_cols) shape as xlsx_preview_rows, for a
    generated .csv file. `total_cols` is the widest row actually seen (a CSV
    has no fixed column count the way a worksheet does). Raises ValueError
    on a decode failure — malformed UTF-8 is the only realistic failure mode
    for a plain-text format."""
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError as err:
        raise ValueError("not valid UTF-8 text") from err
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)
    total_rows = len(all_rows)
    total_cols = max((len(row) for row in all_rows), default=0)
    shown = all_rows[:_PREVIEW_MAX_ROWS]
    rows = [row[:_PREVIEW_MAX_COLS] for row in shown]
    return rows, total_rows, total_cols


def _cell_text(value: object) -> str:
    return "" if value is None else str(value)


def xlsx_to_text(raw: bytes, filename: str) -> str:
    """Render every sheet in `raw` (an .xlsx workbook) as a bounded text
    table.

    `data_only=True`: a formula cell yields its last-CACHED value — whatever
    Excel itself computed and saved — not a live recalculation, since
    openpyxl cannot evaluate formulas. A workbook that was generated
    programmatically and never opened in Excel has no cached value for a
    formula cell, so it reads as None (rendered here as an empty string),
    not the formula text or an error.

    Raises ValueError (the caller turns this into a 422) for anything
    openpyxl can't parse as a valid .xlsx file — deliberately broad, since
    openpyxl's own failure modes for a corrupt/malformed upload aren't a
    single documented exception type.
    """
    try:
        workbook = load_workbook(BytesIO(raw), data_only=True, read_only=True)
    except Exception as err:
        raise ValueError(f"{filename} is not a valid .xlsx file") from err

    try:
        sections: list[str] = []
        for sheet in workbook.worksheets:
            total_rows = sheet.max_row or 0
            total_cols = sheet.max_column or 0
            shown_rows = min(total_rows, _MAX_ROWS_PER_SHEET)
            shown_cols = min(total_cols, _MAX_COLS_PER_SHEET)

            lines = [f"## Sheet: {sheet.title}"]
            if shown_rows > 0 and shown_cols > 0:
                for row in sheet.iter_rows(
                    min_row=1,
                    max_row=shown_rows,
                    max_col=shown_cols,
                    values_only=True,
                ):
                    lines.append("\t".join(_cell_text(v) for v in row))
            if total_rows > _MAX_ROWS_PER_SHEET or total_cols > _MAX_COLS_PER_SHEET:
                lines.append(
                    f"[truncated: sheet has {total_rows} rows x {total_cols} "
                    f"columns; showing the first {shown_rows} rows x "
                    f"{shown_cols} columns]"
                )
            sections.append("\n".join(lines))
    finally:
        workbook.close()

    if not sections:
        return f"{filename} contains no sheets."
    return f"Extracted from {filename}:\n\n" + "\n\n".join(sections)


def resolve_xlsx_attachments(
    files: list[FileAttachment] | None,
) -> list[FileAttachment] | None:
    """Replace every .xlsx entry in `files` with a plain-text FileAttachment
    holding its extracted table text, in place — a non-.xlsx entry (PDF,
    already-plain-text) passes through untouched. Returns `files` unchanged
    (same object) when there's nothing to convert, so a caller can call this
    unconditionally without a branch for the common no-spreadsheet case.

    Raises `HTTPException(422)` for a malformed base64 payload or a
    corrupt/invalid .xlsx file — a v1 scope decision (see this module's own
    docstring and CHANGELOG): fail the whole request cleanly rather than
    guessing at a partial extraction.
    """
    if not files or not any(f.data.startswith(_XLSX_DATA_URL_PREFIX) for f in files):
        return files

    resolved: list[FileAttachment] = []
    for file in files:
        if not file.data.startswith(_XLSX_DATA_URL_PREFIX):
            resolved.append(file)
            continue
        b64 = file.data[len(_XLSX_DATA_URL_PREFIX) :]
        try:
            raw = base64.b64decode(b64, validate=True)
        except (binascii.Error, ValueError) as err:
            raise HTTPException(
                status_code=422, detail=f"{file.filename} is not valid base64"
            ) from err
        try:
            text = xlsx_to_text(raw, file.filename)
        except ValueError as err:
            logger.warning("spreadsheet.parse_failed filename=%s", file.filename)
            raise HTTPException(status_code=422, detail=str(err)) from err
        encoded = base64.b64encode(text.encode("utf-8")).decode("ascii")
        resolved.append(
            FileAttachment(
                filename=file.filename, data=f"data:text/plain;base64,{encoded}"
            )
        )
    return resolved
