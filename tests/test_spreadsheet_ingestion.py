"""Spreadsheet (.xlsx) input: server-side conversion to a bounded text
table, folded into the same document-attachment path a PDF/plain-text file
already goes through — see app/spreadsheet_ingestion.py.
"""

from __future__ import annotations

import base64
from io import BytesIO

import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import Workbook

import app.routers.messages
from app import spreadsheet_ingestion
from app.schemas import AskRequest, AskResponse, FileAttachment
from app.spreadsheet_ingestion import (
    _MAX_COLS_PER_SHEET,
    _MAX_ROWS_PER_SHEET,
    resolve_xlsx_attachments,
    xlsx_to_text,
)

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _workbook_bytes(build) -> bytes:
    """`build(workbook)` mutates a fresh Workbook in place; returns its
    saved .xlsx bytes."""
    wb = Workbook()
    build(wb)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_data_url(raw: bytes) -> str:
    return f"data:{_XLSX_MIME};base64," + base64.b64encode(raw).decode()


# --- xlsx_to_text (module-level) ---------------------------------------------


def test_single_sheet_renders_as_a_tab_separated_table() -> None:
    def build(wb):
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["name", "score"])
        ws.append(["alice", 10])
        ws.append(["bob", 20])

    text = xlsx_to_text(_workbook_bytes(build), "scores.xlsx")

    assert "Extracted from scores.xlsx:" in text
    assert "## Sheet: Sheet1" in text
    assert "name\tscore" in text
    assert "alice\t10" in text
    assert "bob\t20" in text
    assert "truncated" not in text


def test_multi_sheet_workbook_includes_every_sheet() -> None:
    def build(wb):
        ws1 = wb.active
        ws1.title = "Revenue"
        ws1.append(["Q1", 100])
        ws2 = wb.create_sheet("Costs")
        ws2.append(["Q1", 40])

    text = xlsx_to_text(_workbook_bytes(build), "budget.xlsx")

    assert "## Sheet: Revenue" in text
    assert "## Sheet: Costs" in text
    assert "Q1\t100" in text
    assert "Q1\t40" in text


def test_truncation_note_appears_only_when_caps_are_exceeded() -> None:
    def small(wb):
        ws = wb.active
        ws.append(["a", "b"])

    assert "truncated" not in xlsx_to_text(_workbook_bytes(small), "small.xlsx")

    def too_many_rows(wb):
        ws = wb.active
        for i in range(_MAX_ROWS_PER_SHEET + 5):
            ws.append([i])

    text = xlsx_to_text(_workbook_bytes(too_many_rows), "tall.xlsx")
    assert f"has {_MAX_ROWS_PER_SHEET + 5} rows" in text
    assert f"showing the first {_MAX_ROWS_PER_SHEET} rows" in text

    def too_many_cols(wb):
        ws = wb.active
        ws.append(list(range(_MAX_COLS_PER_SHEET + 3)))

    text = xlsx_to_text(_workbook_bytes(too_many_cols), "wide.xlsx")
    assert f"{_MAX_COLS_PER_SHEET + 3} columns" in text
    # Only 1 row was ever written, so the row count itself isn't capped --
    # only the column count is.
    assert f"showing the first 1 rows x {_MAX_COLS_PER_SHEET} columns" in text


def test_a_row_within_the_cap_is_never_truncated_at_the_exact_boundary() -> None:
    def exactly_at_cap(wb):
        ws = wb.active
        for i in range(_MAX_ROWS_PER_SHEET):
            ws.append([i])

    text = xlsx_to_text(_workbook_bytes(exactly_at_cap), "exact.xlsx")
    assert "truncated" not in text


def test_formula_cell_with_no_cached_value_renders_as_empty() -> None:
    """openpyxl can't evaluate formulas -- a workbook built (and saved) by
    openpyxl itself, never opened in Excel, has no cached result for a
    formula cell. data_only=True then reads it as None, not the formula
    text or an error -- this is the documented caveat, pinned here."""

    def build(wb):
        ws = wb.active
        ws.append(["input", "computed"])
        ws["A2"] = 5
        ws["B2"] = "=A2*2"

    text = xlsx_to_text(_workbook_bytes(build), "formulas.xlsx")

    lines = [line for line in text.splitlines() if line.startswith("5")]
    assert lines == ["5\t"]  # the formula cell renders as an empty string


def test_empty_workbook_reports_no_sheets_content() -> None:
    def build(wb):
        ws = wb.active
        ws.title = "Empty"

    text = xlsx_to_text(_workbook_bytes(build), "empty.xlsx")
    assert "## Sheet: Empty" in text


def test_malformed_bytes_raise_value_error() -> None:
    with pytest.raises(ValueError, match="not a valid .xlsx file"):
        xlsx_to_text(b"not an xlsx file at all", "bad.xlsx")


# --- resolve_xlsx_attachments -------------------------------------------------


def test_no_files_returns_the_same_value() -> None:
    assert resolve_xlsx_attachments(None) is None
    files: list[FileAttachment] = []
    assert resolve_xlsx_attachments(files) is files


def test_non_xlsx_files_pass_through_unchanged() -> None:
    files = [FileAttachment(filename="notes.txt", data="data:text/plain;base64,aGk=")]
    assert resolve_xlsx_attachments(files) == files


def test_xlsx_file_is_converted_to_a_plain_text_attachment() -> None:
    def build(wb):
        ws = wb.active
        ws.append(["x", "y"])
        ws.append([1, 2])

    raw = _workbook_bytes(build)
    files = [FileAttachment(filename="data.xlsx", data=_xlsx_data_url(raw))]

    resolved = resolve_xlsx_attachments(files)

    assert resolved is not None and len(resolved) == 1
    assert resolved[0].data.startswith("data:text/plain;base64,")
    decoded = base64.b64decode(resolved[0].data.split(",", 1)[1]).decode()
    assert "x\ty" in decoded
    assert "1\t2" in decoded


def test_xlsx_conversion_preserves_other_attachments_and_their_order() -> None:
    def build(wb):
        ws = wb.active
        ws.append(["a"])

    raw = _workbook_bytes(build)
    files = [
        FileAttachment(filename="before.txt", data="data:text/plain;base64,aGk="),
        FileAttachment(filename="data.xlsx", data=_xlsx_data_url(raw)),
        FileAttachment(filename="after.txt", data="data:text/plain;base64,eW8="),
    ]

    resolved = resolve_xlsx_attachments(files)

    assert resolved is not None
    assert [f.filename for f in resolved] == ["before.txt", "data.xlsx", "after.txt"]
    assert resolved[1].data.startswith("data:text/plain;base64,")


def test_malformed_base64_raises_422() -> None:
    # "QQ=" uses only base64-alphabet characters (so it passes
    # FileAttachment's regex validator at construction time) but has bad
    # padding, so it reaches -- and fails inside -- our own
    # base64.b64decode(..., validate=True) call.
    files = [FileAttachment(filename="bad.xlsx", data=f"data:{_XLSX_MIME};base64,QQ=")]
    with pytest.raises(HTTPException) as exc_info:
        resolve_xlsx_attachments(files)
    assert exc_info.value.status_code == 422


def test_corrupt_xlsx_bytes_raise_422_not_500() -> None:
    corrupt = base64.b64encode(b"this is not really an xlsx file").decode()
    files = [
        FileAttachment(filename="bad.xlsx", data=f"data:{_XLSX_MIME};base64,{corrupt}")
    ]
    with pytest.raises(HTTPException) as exc_info:
        resolve_xlsx_attachments(files)
    assert exc_info.value.status_code == 422


# --- HTTP integration: POST /v1/conversations/{id}/ask ------------------------


def _stub_run_orchestrator(monkeypatch: pytest.MonkeyPatch) -> list[AskRequest]:
    captured: list[AskRequest] = []

    def fake_run_orchestrator(
        req: AskRequest,
        routing_question: str | None = None,
        owner: str | None = None,
        history: str = "",
        cacheable_system: str | None = None,
        anthropic_question: str | None = None,
        context_free: bool = False,
        pre_stage_timings: dict | None = None,
        library_sources: list | None = None,
        forced_category: str | None = None,
    ) -> AskResponse:
        captured.append(req)
        return AskResponse(answer="canned answer", mode_used="auto->fast", notes="n")

    monkeypatch.setattr(app.routers.messages, "run_orchestrator", fake_run_orchestrator)
    return captured


def _create(client: TestClient) -> int:
    return int(client.post("/v1/conversations", json={"title": "t"}).json()["id"])


def test_ask_with_xlsx_converts_before_the_model_ever_sees_it(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    def build(wb):
        ws = wb.active
        ws.append(["quarter", "revenue"])
        ws.append(["Q1", 1000])

    raw = _workbook_bytes(build)
    captured = _stub_run_orchestrator(monkeypatch)
    cid = _create(client)

    res = client.post(
        f"/v1/conversations/{cid}/ask",
        json={
            "question": "summarize this",
            "files": [{"filename": "budget.xlsx", "data": _xlsx_data_url(raw)}],
        },
    )
    assert res.status_code == 200

    assert captured[0].files is not None
    assert captured[0].files[0].data.startswith("data:text/plain;base64,")
    decoded = base64.b64decode(captured[0].files[0].data.split(",", 1)[1]).decode()
    assert "Q1\t1000" in decoded

    messages = client.get(f"/v1/conversations/{cid}/messages").json()
    user_message = next(m for m in messages if m["role"] == "user")
    assert user_message["files"][0]["data"].startswith("data:text/plain;base64,")


def test_ask_with_a_corrupt_xlsx_fails_cleanly_with_422(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    _stub_run_orchestrator(monkeypatch)
    cid = _create(client)
    corrupt = base64.b64encode(b"definitely not a real workbook").decode()

    res = client.post(
        f"/v1/conversations/{cid}/ask",
        json={
            "question": "summarize this",
            "files": [
                {"filename": "bad.xlsx", "data": f"data:{_XLSX_MIME};base64,{corrupt}"}
            ],
        },
    )
    assert res.status_code == 422


def test_regenerate_reuses_the_extracted_table_without_reparsing(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    parse_calls = {"count": 0}
    original = spreadsheet_ingestion.xlsx_to_text

    def counting_xlsx_to_text(raw: bytes, filename: str) -> str:
        parse_calls["count"] += 1
        return original(raw, filename)

    monkeypatch.setattr(spreadsheet_ingestion, "xlsx_to_text", counting_xlsx_to_text)

    def build(wb):
        ws = wb.active
        ws.append(["a", "b"])

    raw = _workbook_bytes(build)
    _stub_run_orchestrator(monkeypatch)
    cid = _create(client)

    client.post(
        f"/v1/conversations/{cid}/ask",
        json={
            "question": "summarize",
            "files": [{"filename": "sheet.xlsx", "data": _xlsx_data_url(raw)}],
        },
    )
    assert parse_calls["count"] == 1

    res = client.post(f"/v1/conversations/{cid}/regenerate", json={})
    assert res.status_code == 200
    # Regenerate has no `files` field at all -- it re-reads the
    # already-persisted (already-converted) user message, so the workbook is
    # never re-parsed.
    assert parse_calls["count"] == 1
